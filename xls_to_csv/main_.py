from PyQt5.QtWidgets import QApplication, QWidget, QDesktopWidget, QHBoxLayout, QFormLayout
from PyQt5.QtWidgets import (
    QLabel,
    QVBoxLayout,
    QLineEdit,
    QPushButton,
    QProgressBar
)
from PyQt5.QtCore import Qt, QTimer, QSize
import style
from pathlib import Path
from openpyxl import load_workbook
import csv


class MainWindow(QWidget):
    height = 200
    width = 600

    def __init__(self):
        super().__init__()

        # windows properties
        self.setGeometry(100, 100, MainWindow.width, MainWindow.height)
        self.setWindowTitle('Excel to CSV')

        # fixing the size
        self.setMaximumSize(MainWindow.width, MainWindow.height)
        self.setMinimumSize(MainWindow.width, MainWindow.height)

        # puts window in center of screen
        current_position = self.frameGeometry()
        screen_geometry = QDesktopWidget().availableGeometry().center()
        current_position.moveCenter(screen_geometry)
        self.move(current_position.topLeft())

        # create, style and show ui
        self.setup_ui()
        self.set_style()
        self.show()

    def setup_ui(self):
        # layout
        main_layout = QVBoxLayout()
        form_layout = QFormLayout()
        progress_layout = QHBoxLayout()
        main_layout.addLayout(form_layout)
        main_layout.addLayout(progress_layout)

        # widgets
        self.excel_path_txt = QLineEdit()
        self.excel_path_txt.setPlaceholderText(r'D:\temp\excel_file_name.xlsx')
        self.csv_path_txt = QLineEdit()
        self.csv_path_txt.setPlaceholderText(r'D:\temp')

        self.convert_btn = QPushButton('Convert')
        self.convert_btn.clicked.connect(self.convert)

        self.output_lbl = QLabel('sheet out of sheets')

        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(False)

        # add widget to layout
        form_layout.addRow('Excel File Path: ', self.excel_path_txt)
        form_layout.addRow('Save to: ', self.csv_path_txt)
        form_layout.addRow('', self.convert_btn)
        progress_layout.addWidget(self.output_lbl)
        progress_layout.addWidget(self.progress_bar)

        self.setLayout(main_layout)

    def convert(self):
        count_progress_bar = 0

        # get path
        source_path = self.excel_path_txt.text().encode('unicode_escape').decode().replace('\\\\', '\\')
        destination_path = self.csv_path_txt.text().encode('unicode_escape').decode().replace('\\\\', '\\')
        print(source_path)
        print(destination_path)

        # load workbook
        wb = load_workbook(filename=source_path, data_only=True)
        sheets = wb.sheetnames
        sheets_count = len(sheets)

        # set progress bar range
        self.progress_bar.setMaximum(sheets_count)

        for sheet_name in sheets:
            # activate sheets
            wb.active = wb[sheet_name]

            # read data and save as csv
            path = Path(destination_path, sheet_name).with_suffix('.csv')
            with open(path, 'w', newline='') as f:
                csv_writer = csv.writer(f)

                for row in wb.active.rows:
                    data = [cell.value for cell in row]
                    csv_writer.writerow(data)

            # update progress bar
            count_progress_bar += 1
            self.output_lbl.setText(f'{count_progress_bar} out of {sheets_count}')
            self.progress_bar.setValue(count_progress_bar)

    def set_style(self):
        # text box
        self.excel_path_txt.setStyleSheet(style.LINE_EDIT_STYLE)
        self.csv_path_txt.setStyleSheet(style.LINE_EDIT_STYLE)

        # button
        self.convert_btn.setStyleSheet(style.CONVERT_BUTTON_STYLE)

        # label
        self.output_lbl.setStyleSheet(style.LABEL_STYLE)

        # progress bar
        self.progress_bar.setStyleSheet(style.progress_bar_style)


def main():
    app = QApplication([])
    ui = MainWindow()
    app.exec_()


if __name__ == '__main__':
    main()
